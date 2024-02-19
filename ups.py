#import new token method
from auth import new_token, test_bearer_token
import requests
from xl import get_num_cols, get_num_rows, copy_row
from response import trackPackage
from openpyxl import workbook, load_workbook, worksheet
from openpyxl.utils import get_column_letter

file = open("./auth/bearer.txt", "r")
bear_tok = file.read()
query = {
  "locale": "en_US",
  "returnSignature": "false",
  "returnMilestones": "false"
}

headers = {
  "transId": "absupstracking01",
  "transactionSrc": "tracking",
  "Authorization": f"Bearer {bear_tok}"
}

#CALLS TRACKING API AND RETURNS INFO ON THE PASSED IN TRACKING NUMBER
def call_api(tracking_number): 

  #IF BEARER TOKEN IS EXPIRED
  if(test_bearer_token(bear_tok) == 401):
    #GET NEW BEARER TOKEN
    headers['Authorization'] = f'Bearer str({new_token()})'
  
  #URL APPENDS PASSED TRACKING NUMBER STRING TO URL API ENDPOINT
  url = "	https://onlinetools.ups.com/api/track/v1/details/" + tracking_number

  #TRY CALLING TRACKING API ENDPOINT
  try:
    response = requests.get(url, headers=headers, params=query)
    data = response.json()

    #IF RESPONSE IS GOOD, PASS DATA INTO TRACKPACKAGE METHOD
    if(response.status_code == 200):
      list_for_xl = trackPackage(data)
      return list_for_xl
    
    else:
      return 'fail'

  except Exception as e: 
    print(f"fail! {e}")
   


#OPENS WORKBOOK, CALLS API, AND UPDATES WORKBOOK APPROPIRATELY
if __name__ == '__main__':

  #LOADS THE WORKBOOK AND 2 SHEETS INSIDE OF THE WORKBOOK
  wb = load_workbook(r"C:\Users\MForet\OneDrive - Alternative Behavior Strategies, LLC\UPS Tracking.xlsx")
  delivered_rows = []
  ws = wb['Tracking']
  ws2 = wb['Completed']

  #LOOPS THROUGH THE ROWS THAT HAVE TRACKING INFO
  for row in range(2,get_num_rows(ws)+1):

    #IF THE PACKAGE WAS PREVIOUSLY MARKED AS DELIVERED, WE COPY THE ROW AND ADD IT TO OUR DELIVERED ROWS LIST ALONG WITH THE ROW NUMBER
    if(ws['C'+str(row)].value == 'Delivered'):
      delivered_rows.append([row, copy_row(ws, row)])

    #IF THE PACKAGE WAS NOT PREVIOUSLY MARKED AS DELIVERED
    else:

      #CALL UPS API WHICH RETURNS A LIST WITH THE TRACKING INFO
      list = call_api(ws['A'+str(row)].value)

      #IF LIST IS NOT EMPTY WE LOOP THROUGH THE LIST AND UPDATE THE EXCEL ROW
      if(list != 'fail'):
        count = 2

        for key in list:
          ws[get_column_letter(count)+str(row)] = list[key]
          count += 1

      #IF LIST IS EMPTY, THERE WAS AN ISSUE RETREVIING DATA AND THE PROGRAM ENDS.
      else:
        print('list is empty. Failed to get data from UPS.')
        exit(1)

  
  #IF THERE ARE ANY ITEMS THAT WERE MARKED AS DELIVERED LAST TIME THE SCRIPT RAN
    #IT WILL MOVE THE ROW TO THE COMPLETED TAB
  if(delivered_rows):
    num_deleted=0
    for item in delivered_rows:
      index_to_delete=item[0]
      ws.delete_rows(int(item[0])-num_deleted)
      ws2.append(item[1])
      num_deleted+=1
      
  #SAVE WORKBOOK
  wb.save(r"C:\Users\MForet\OneDrive - Alternative Behavior Strategies, LLC\UPS Tracking.xlsx")
