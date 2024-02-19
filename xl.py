#!/usr/bin/python3.11
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

#RETURNS NUMBER OF FILLED COLUMNS IN WORKSHEET
def get_num_rows(ws, col="A"):
    count = 1
    x=''

    while(x != "None"):

        x = str(ws[col+str(count)].value)
        count += 1

    return count-2

#RETURNS NUMBER OF FILLED COLUMNS IN WORKSHEET
def get_num_cols(ws, row):
    count = 1
    x=''

    while(x != "None"):

        char = get_column_letter(count)
        x = str(ws[char+str(row)].value)
        count += 1
    
    return count-2

#RETURNS LIST OF ROW VALUES FOR SPECIFIED ROW
def copy_row(ws, row):
    
    array = []

    for col in range(1, get_num_cols(ws, row)+1):

        char = get_column_letter(col)
        array.append(ws[char+str(row)].value)
    
    return array